from openpyxl import load_workbook
import openpyxl
import xlwt
from datetime import datetime

import os

def contador():
        FILE_PATH = 'docu.xlsx'
        SHEET = 'Hoja1'
        workbook = load_workbook(FILE_PATH, read_only=True)
        sheet = workbook[SHEET]
        i=0
        for row in sheet.iter_rows():

           i = i+1

        print(i)
        return str(i)

def contador2():
        FILE_PATH = 'docu.xlsx'
        SHEET = 'Hoja2'
        workbook = load_workbook(FILE_PATH, read_only=True)
        sheet2 = workbook[SHEET]
        i=0
        for row in sheet2.iter_rows():

           i = i+1

        print(i)
        return str(i)

def menu():

    os.system('cls')
    print("SELECCIONES UNA OPCION:  ")
    print("1 - REGRISTRAR PROVEEDORES")
    print("2 - REGISTRAR ARTICULOS ELECTRONICOS")
    print("3 - LISTAR PROVEEDORES")
    print("4 - LISTAR ARTICULOS ELECTRONICOS")
    print("9 - SALIR DEL PROGRAMA")
    

numero = 0
while(numero != 9):

    menu()

    numero = input("SELLECIONE UN VALOR:  ")

    if numero == "1":
        os.system('cls')
        
        doc = openpyxl.load_workbook('docu.xlsx')
        hoja = doc.get_sheet_by_name('Hoja1')
     
        i = contador()
        
        celda_A ="A"+i
        celda_B ="B"+i
        celda_C ="C"+i
        celda_D ="D"+i
        celda_E ="E"+i
        celda_F ="F"+i

        

        hoja[celda_A] = input("Ingrese ID:  ")
        hoja[celda_B] = input("Ingrese Nombre:  ")
        hoja[celda_C] = input("Ingrese País:  ")
        hoja[celda_D] = input("Ingrese Dirección:  ")
        hoja[celda_E] = input("Ingrese Telefono:  ")
        hoja[celda_F] = input("Ingrese Email:  ")
        
        doc.save("docu.xlsx")
        os.system('pause')

    elif numero == "2":
        os.system('cls')
        
        doc = openpyxl.load_workbook('docu.xlsx')
        hoja = doc.get_sheet_by_name('Hoja2')
     
        i = contador2()
        
        celda_A ="A"+i
        celda_B ="B"+i
        celda_C ="C"+i
        celda_D ="D"+i
        celda_E ="E"+i
        celda_F ="F"+i
        celda_G ="G"+i
        celda_H ="H"+i
        celda_I ="I"+i

        

        hoja[celda_A] = input("Ingrese ID:  ")
        hoja[celda_B] = input("Ingrese Nombre:  ")
        hoja[celda_C] = input("Ingrese Categoria:  ")
        hoja[celda_D] = input("Ingrese Marca:  ")
        hoja[celda_E] = input("Ingrese Modelo:  ")
        hoja[celda_F] = input("Ingrese Continente:  ")
        hoja[celda_G] = input("Ingrese Precio:  ")
        hoja[celda_H] = input("Ingrese Cantidad:  ")
        hoja[celda_I] = input("Ingrese proveedor:  ")
        
        doc.save("docu.xlsx")
        os.system('pause')

    elif numero == "3":
            
        os.system('cls')
        FILE_PATH = 'docu.xlsx'
        SHEET = 'Hoja1'
        workbook = load_workbook(FILE_PATH, read_only=True)
        sheet = workbook[SHEET]
       
        for row in sheet.iter_rows():
            
            print(row[1].value)
           
        
        
        os.system('pause')
    elif numero == "4":
        os.system('cls')
        FILE_PATH = 'docu.xlsx'
        SHEET = 'Hoja2'
        workbook = load_workbook(FILE_PATH, read_only=True)
        sheet = workbook[SHEET]

        for row in sheet.iter_rows():
            
            print(row[1].value)

        os.system('pause')    
    
    elif numero == "9":
            break
        
    else:

            print("")
            input("opcion incorrecta... Prueba otra vez")
        